import pandas as pd

from flask import Blueprint, render_template, request, redirect, url_for
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook

# Definindo o Blueprint para o projeto 4
projeto4_blueprint = Blueprint('projeto4', __name__, template_folder='templates', static_folder='static')

# Caminho do arquivo Excel
EXCEL_FILE_PATH = r'C:\Users\victo\OneDrive\Documentos\sistema_unificado\dados.xlsx'


# Função para traduzir o nome dos meses para português
def traduzir_mes(mes_ingles):
    meses_pt_br = {
        'January': 'Janeiro', 'February': 'Fevereiro', 'March': 'Março',
        'April': 'Abril', 'May': 'Maio', 'June': 'Junho',
        'July': 'Julho', 'August': 'Agosto', 'September': 'Setembro',
        'October': 'Outubro', 'November': 'Novembro', 'December': 'Dezembro'
    }
    return meses_pt_br.get(mes_ingles, mes_ingles)

# Rotas do Blueprint
@projeto4_blueprint.route('/')
def index():
    return render_template('projeto4.html')

@projeto4_blueprint.route('/registrar', methods=['POST'])
def registrar():
    # Coleta os dados do formulário
    data = {
        "Equipamento": request.form['equipment'],
        "Data": datetime.strptime(request.form['date'], '%Y-%m-%d').strftime('%d/%m/%Y'),
        "Hora": request.form['time'],
        "Responsável": request.form['responsible'],
        "Temperatura Atual": request.form.get('current-temp', 'N/A'),
        "Temperatura Máxima": request.form.get('max-temp', 'N/A'),
        "Temperatura Mínima": request.form.get('min-temp', 'N/A'),
        "Observações": request.form.get('observations', 'Nenhum'),
        "Observações Texto": request.form.get('observations-text', '')
    }
    
    sheet_name = request.form['equipment'].strip()  # Remove espaços em branco adicionais

    try:
        # Cria um novo arquivo se não existir
        if not os.path.exists(EXCEL_FILE_PATH):
            book = Workbook()
            book.save(EXCEL_FILE_PATH)
        
        # Carrega o arquivo existente
        book = load_workbook(EXCEL_FILE_PATH)
        
        # Se a aba não existir, cria uma nova
        if sheet_name not in book.sheetnames:
            sheet = book.create_sheet(title=sheet_name)
            # Adiciona cabeçalhos na nova aba
            headers = list(data.keys())
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)
        else:
            sheet = book[sheet_name]
        
        # Encontra a próxima linha vazia
        next_row = sheet.max_row + 1
        
        # Adiciona o novo registro na próxima linha
        for col_num, header in enumerate(data.keys(), 1):
            sheet.cell(row=next_row, column=col_num, value=data[header])
        
        # Salva as mudanças
        book.save(EXCEL_FILE_PATH)

    except Exception as e:
        print(f"Erro ao manipular o arquivo Excel: {e}")
        return f"Ocorreu um erro ao processar o arquivo: {e}", 500

    return redirect(url_for('projeto4.index'))

@projeto4_blueprint.route('/relatorio')
def relatorio():
    try:
        # Verifica se o arquivo existe
        if not os.path.exists(EXCEL_FILE_PATH):
            return "O arquivo de dados não existe.", 404

        # Carrega todos os dados das planilhas
        df = pd.read_excel(EXCEL_FILE_PATH, engine='openpyxl', sheet_name=None)
        instrument_reports = []

        for sheet_name, data in df.items():
            if 'Data' not in data.columns:
                print(f"Coluna 'Data' não encontrada na aba '{sheet_name}'.")
                continue

            # Converte a data para o formato datetime
            try:
                data['Data'] = pd.to_datetime(data['Data'], format='%d/%m/%Y', errors='coerce')
                data = data.dropna(subset=['Data'])  # Remove entradas com data inválida
                # Formata a data para exibição
                data['Data'] = data['Data'].dt.strftime('%d/%m/%Y')
            except Exception as e:
                print(f"Erro ao formatar a data na aba '{sheet_name}': {e}")
                continue

            # Ordena os dados por data
            data = data.sort_values(by='Data')

            # Formata temperaturas
            for col in ['Temperatura Atual', 'Temperatura Máxima', 'Temperatura Mínima']:
                if col in data.columns:
                    # Converte para numérico e formata
                    data[col] = pd.to_numeric(data[col], errors='coerce')
                    data[col] = data[col].apply(lambda x: f"{x:.1f} ºC" if pd.notnull(x) else "N/A")

            # Adiciona o relatório à lista de relatórios
            for month_year, group in data.groupby(data['Data'].str[3:]):  # Agrupa por mês e ano
                mes_ingles = datetime.strptime(month_year, '%m/%Y').strftime('%B')
                mes_portugues = traduzir_mes(mes_ingles)
                report_name = f"Termômetro - {sheet_name}"  # Nome do equipamento com prefixo
                instrument_reports.append({
                    'sheet_name': report_name,
                    'month_year': f"{mes_portugues} {datetime.strptime(month_year, '%m/%Y').year}",
                    'html_table': group.to_html(classes='table table-striped table-bordered text-center', index=False, header=True),
                    'month_year_full': f"{mes_portugues}/{datetime.strptime(month_year, '%m/%Y').year}"
                })

        if not instrument_reports:
            return "Nenhum dado encontrado para exibição.", 404

        return render_template('relatorio_temperatura.html', instrument_reports=instrument_reports)
    except Exception as e:
        print(f"Erro ao ler o arquivo Excel para o relatório: {e}")
        return f"Ocorreu um erro ao processar o relatório: {e}", 500
