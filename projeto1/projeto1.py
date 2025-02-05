from flask import Blueprint, render_template, request, jsonify, url_for, redirect
from flask_login import login_required
from openpyxl import load_workbook
import os
from datetime import datetime

# Criar o Blueprint
projeto1_blueprint = Blueprint('projeto1', __name__, template_folder='templates', static_folder='static')

# Caminho da planilha de controle de pH
PLANILHA_PATH = r"Z:\Controles do Setor de Meios de Cultura\Cartas controle\Carta Controle de Exatidão de pH 4,00 e 7,00.xlsx"

# Rota para o formulário
@projeto1_blueprint.route("/", methods=["GET", "POST"])
@login_required
def formulario():
    if request.method == "POST":
        equipamento = request.form["equipamento"]
        padrao = request.form["padrao"]
        valor_lido = request.form["valor_lido"]
        data = request.form["data"]
        analista = request.form["analista"]

        # Formatar a data para o formato dd/mm/aaaa
        data_formatada = datetime.strptime(data, "%Y-%m-%d").strftime("%d/%m/%Y")

        # Nome da aba baseado no equipamento e no padrão
        aba = f"{equipamento} Padrão {padrao.replace('.', ',')}"

        resultado = inserir_dados_na_planilha(aba, data_formatada, padrao, valor_lido, analista)

        if "Erro" in resultado:
            return jsonify({"status": "error", "message": resultado})
        else:
            return jsonify({
                "status": "success",
                "message": resultado,
                "equipamento": equipamento,
                "padrao": padrao,
                "valor_lido": valor_lido,
                "data": data_formatada,
                "analista": analista
            })

    return render_template("index.html")  # Certifique-se de que index.html está em 'projeto1/templates'

def inserir_dados_na_planilha(aba, data, padrao, valor_lido, analista):
    if not os.path.exists(PLANILHA_PATH):
        return "Erro: Planilha não encontrada."

    try:
        wb = load_workbook(PLANILHA_PATH)
    except Exception as e:
        return f"Erro ao abrir a planilha: {e}"

    if aba not in wb.sheetnames:
        return f"Erro: Aba '{aba}' não encontrada na planilha."

    ws = wb[aba]

    if ws["D16"].value != "Resultado":
        return "Erro: A célula D16 não contém o cabeçalho 'Resultado'."

    next_row = 17
    while ws[f"B{next_row}"].value is not None:
        next_row += 1

    ws[f"B{next_row}"] = data
    ws[f"C{next_row}"] = float(padrao)
    ws[f"D{next_row}"] = float(valor_lido)
    ws[f"N{next_row}"] = analista

    try:
        wb.save(PLANILHA_PATH)
        return "Dados registrados com sucesso!"
    except Exception as e:
        return f"Erro ao salvar a planilha: {e}"
