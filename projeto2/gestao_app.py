import pandas as pd
import openpyxl
import os

# Caminho do arquivo de origem (controle de qualidade)
file_path = r'C:\Users\victo\OneDrive\Área de Trabalho\Controle de Qualidade.xlsx'

# Caminho da planilha de destino (relatório de produção)
output_file = r'C:\Users\victo\OneDrive\Documentos\Controle de estoque\projeto_gestao_estoque\relatorio_producao.xlsx'

# Verificar se o arquivo de origem existe
if not os.path.exists(file_path):
    print(f"Arquivo de origem não encontrado no caminho: {file_path}")
else:
    try:
        # Nome da aba a ser carregada da planilha de origem
        sheet_name = 'ENTRADA | SAÍDA'
        
        # Tentar carregar os dados da aba especificada sem cabeçalho
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)  # Não usar cabeçalho na leitura inicial
        
        # Mostrar as primeiras linhas para verificar a leitura
        print(f"Primeiras linhas carregadas:\n{df.head()}")
        
        # Verificar o número de linhas e colunas
        print(f"Informações sobre a planilha: {df.shape}")
        
        # Caso o DataFrame tenha mais de uma linha, tentar ajustar a leitura
        if df.shape[0] > 1:  # Verificar se existem mais de uma linha
            # Ajustar para pegar o cabeçalho, se necessário, e carregar novamente com base na estrutura
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=4)  # O cabeçalho está na linha 5 (índice 4)
            print(f"Primeiras linhas após o ajuste de cabeçalho:\n{df.head()}")
            
            # Verificar as colunas carregadas para garantir que o nome está correto
            print(f"Nomes das colunas carregadas: {df.columns.tolist()}")
            
            # Verificar se as colunas de interesse existem
            if 'DATA' in df.columns and 'MEIO' in df.columns and 'QUANTITATIVO' in df.columns and 'TIPO' in df.columns:
                # Convertendo a coluna 'DATA' para o tipo datetime
                df['DATA'] = pd.to_datetime(df['DATA'], errors='coerce')
                
                # Exibindo as primeiras linhas após a conversão
                print(f"Primeiras linhas com DATA convertida:\n{df.head()}")
                
                # Filtrar apenas as entradas (descartar 'SAÍDA')
                df = df[df['TIPO'] == 'ENTRADA']
                
                # Processando dados de todos os anos a partir de 2023
                df['ANO'] = df['DATA'].dt.year  # Criar coluna de ano
                
                # Filtrando os dados para os anos a partir de 2023
                df_years = df[df['ANO'] >= 2023]
                
                # Convertendo a coluna 'QUANTITATIVO' para numérico, forçando erros para NaN
                df_years['QUANTITATIVO'] = pd.to_numeric(df_years['QUANTITATIVO'], errors='coerce')
                
                # Garantindo que estamos trabalhando com uma cópia do DataFrame para evitar warning
                df_years = df_years.copy()  
                
                # Adicionando uma coluna 'Mês' com o número do mês
                df_years['Mês'] = df_years['DATA'].dt.month

                # Produção mensal e anual por MEIO
                df_monthly = df_years.groupby(['ANO', 'Mês', 'MEIO'])['QUANTITATIVO'].sum().reset_index()
                df_yearly = df_years.groupby(['ANO', 'MEIO'])['QUANTITATIVO'].sum().reset_index()

                # Exibindo os resultados para checagem
                print(f"Produção mensal por meio:\n{df_monthly}")
                print(f"Produção anual por meio:\n{df_yearly}")
                
                # Abrir a planilha de destino com openpyxl para preservar a formatação
                wb_destino = openpyxl.load_workbook(output_file)

                # Se a planilha de destino não existir, cria uma nova
                if 'Produção Mensal' not in wb_destino.sheetnames:
                    wb_destino.create_sheet('Produção Mensal')
                if 'Produção Anual' not in wb_destino.sheetnames:
                    wb_destino.create_sheet('Produção Anual')
                
                # Acessar as abas de destino
                sheet_monthly = wb_destino['Produção Mensal']
                sheet_yearly = wb_destino['Produção Anual']
                
                # Atualizando os dados na planilha "Produção Mensal" sem apagar a formatação
                for index, row in df_monthly.iterrows():
                    # Encontrar a linha na planilha onde os dados precisam ser inseridos
                    for i, excel_row in enumerate(sheet_monthly.iter_rows(min_row=2, max_row=sheet_monthly.max_row)):
                        if excel_row[0].value == row['ANO'] and excel_row[1].value == row['Mês'] and excel_row[2].value == row['MEIO']:
                            # Atualizar o valor na célula correspondente
                            sheet_monthly.cell(row=i+2, column=4, value=row['QUANTITATIVO'])
                            break
                    else:
                        # Se não encontrar a linha, adicionar uma nova
                        sheet_monthly.append([row['ANO'], row['Mês'], row['MEIO'], row['QUANTITATIVO']])

                # Atualizando os dados na planilha "Produção Anual" sem apagar a formatação
                for index, row in df_yearly.iterrows():
                    # Encontrar a linha na planilha onde os dados precisam ser inseridos
                    for i, excel_row in enumerate(sheet_yearly.iter_rows(min_row=2, max_row=sheet_yearly.max_row)):
                        if excel_row[0].value == row['ANO'] and excel_row[1].value == row['MEIO']:
                            # Atualizar o valor na célula correspondente
                            sheet_yearly.cell(row=i+2, column=3, value=row['QUANTITATIVO'])
                            break
                    else:
                        # Se não encontrar a linha, adicionar uma nova
                        sheet_yearly.append([row['ANO'], row['MEIO'], row['QUANTITATIVO']])

                # Salvar o arquivo de destino com os dados atualizados, mantendo a formatação intacta
                wb_destino.save(output_file)
                
                print(f"Relatório atualizado e salvo em: {output_file}")
            else:
                print("As colunas 'DATA', 'MEIO', 'QUANTITATIVO' ou 'TIPO' não foram encontradas no arquivo.")
        else:
            print("O arquivo contém apenas uma ou duas linhas, e não é possível carregar os dados corretamente.")
    
    except Exception as e:
        print(f"Ocorreu um erro ao processar o arquivo: {e}")
